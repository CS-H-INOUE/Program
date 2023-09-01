import openpyxl
import os

# 有効な名前のリストを定義します（最初の文字が A, B, C, D であるもの）
valid_names = ["井上幸次", "水谷", "大橋", "井出"]


def copy_first_two_rows(source_sheet, output_sheet):
    for row in source_sheet.iter_rows(min_row=1, max_row=2):
        output_sheet.append([cell.value for cell in row])

def copy_products(excel_sheet, output_sheet, start_row, max_row):
    for row_idx in range(start_row, max_row + 1):
        if excel_sheet.cell(row=row_idx, column=1).value is None:
            break
        product = [excel_sheet.cell(row=row_idx, column=col).value for col in range(1, excel_sheet.max_column + 1)]
        output_sheet.append(product)

def separate_sheets_by_names(excel_file_path, output_path):
    # エクセルファイルを開く
    wb = openpyxl.load_workbook(excel_file_path)
    sheet = wb.active

    current_row = 1
    max_row = sheet.max_row

    # 新しいワークブックを作成して、商品データを格納します
    output_wb = openpyxl.Workbook()
    output_wb.create_sheet(title="Sheet")  # デフォルトのシートを作成

    # current_rowを3に設定して処理開始
    current_row = 3

    while current_row <= max_row:
        # 名前のセルの値を取得します
        name = sheet.cell(row=current_row, column=1).value
        
        # デバッグ用に取得した名前を表示
        print(f"Processing name: {name}")

        # 名前が有効な名前リストに含まれているかチェックします
        if name is not None and any(name.startswith(valid_name) for valid_name in valid_names):
            # 名前ごとのシートを作成します
            new_sheet = output_wb.create_sheet(title=name or "Unknown")
            
            # 1行目と2行目のデータをコピーします
            copy_first_two_rows(sheet, new_sheet)
            
            # 3行目以降の商品データをコピーします
            copy_products(sheet, new_sheet, current_row, min(current_row + 40, max_row))
            
            # 処理が完了したら次の行へ移動します
            current_row += 40

        # 名前が無効な場合は次の行へ移動します
        else:
            current_row += 1

    # デフォルトのシートを削除して、ファイルを保存します
    output_wb.remove(output_wb.active)
    output_wb.save(output_path)
    output_wb.close()

def main():
    # エクセルファイルと出力パスを指定します
    excel_file_path = "file.xlsx"
    output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output.xlsx")

    # 商品データをコピーして保存します
    separate_sheets_by_names(excel_file_path, output_path)

    # file2.xlsxの処理
    excel_file_path2 = "file2.xlsx"
    output_path2 = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output2.xlsx")
    separate_sheets_by_names(excel_file_path2, output_path2)

if __name__ == "__main__":
    main()