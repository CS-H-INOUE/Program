import openpyxl
import os

# 有効な名前のリストを定義します（最初の文字が A, B, C, D であるもの）
valid_names = ["井上幸次", "水谷", "大橋", "井出"]


def copy_products_to_single_file(excel_file_path, output_path):
    # エクセルファイルを開く
    wb = openpyxl.load_workbook(excel_file_path)
    sheet = wb.active

    current_row = 1
    max_row = sheet.max_row

    # 新しいワークブックを作成して、商品データを格納します
    output_wb = openpyxl.Workbook()

    while current_row <= max_row:
        # 名前のセルの値を取得します
        name = sheet.cell(row=current_row, column=1).value
        
        # デバッグ用に取得した名前を表示
        print(f"Processing name: {name}")

        # 名前が有効な名前リストに含まれているかチェックします
        if name is not None and any(name.startswith(valid_name) for valid_name in valid_names):
            products = []
            for _ in range(40):  # 各名前に対して40行の商品を取得
                current_row += 1
                if current_row > max_row:
                    break
                product = []
                for col in range(1, sheet.max_column + 1):  # すべての列を取得
                    product.append(sheet.cell(row=current_row, column=col).value)
                products.append(product)

            if products:
                # 名前をタイトルとして新しいシートを作成します
                new_sheet = output_wb.create_sheet(title=name or "Unknown")

                # 商品データを新しいシートに書き込みます
                for row_idx, product in enumerate(products, start=1):
                    for col_idx, value in enumerate(product, start=1):
                        new_sheet.cell(row=row_idx, column=col_idx, value=value)

        # 空欄の名前の場合はループを終了します
        else:
            current_row += 1

    # デフォルトのシートを削除して、ファイルを保存します
    output_wb.remove(output_wb.active)
    output_wb.save(output_path)
    output_wb.close()

def main():
    # エクセルファイルと出力パスを指定します
    excel_file_path = "file.xlsx"
    output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output.xlsx")

    # 商品データをコピーして保存します
    copy_products_to_single_file(excel_file_path, output_path)

if __name__ == "__main__":
    main()