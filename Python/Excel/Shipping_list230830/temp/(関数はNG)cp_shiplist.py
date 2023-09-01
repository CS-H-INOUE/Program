'''
あなたは、プロのコーダーです。
以下のControlとInputから、プログラムを出力してください。
なおInputはない場合は、Noneとします。

###Control
プログラムはpython3でお願いいたします。

Excelファイルを読み込みます。
なおパスは ./現在の年/{当月(2ケタ).xlsx}となります。
 （ex. 2023年8月であれば、./2023/08.xlsxを参照）

その中の"当月""月"(例. 8月なら、8月、12月なら12月)シートのC列が当日のものと一致する行を
新規ファイルのcopyというシートを作成し、その中にコピーしてください。

一致するためのロジックは
エクセルの入力形式が、{月}/{日}となっています。（例. 8/1, 12/31）
したがって、当日の日付を取得し、その日付と一致する行を全て抽出してください。

ペーストは値として処理してください。（元の関数はコピーしないでください。例. vlookup -> 参照された値をいれる、例えば3など）


なおプログラムの可用性やメンテ性などを考慮したわかりやすいコードでコメントも追記しながら記述してください。


→追記
NGだったため、以下を追加質問

vlookなどの関数がそのままペーストされています。
これを実際にvlookで抽出した値（例えば3など）としてコピーしてほしいです。


###Input
None
'''

import os
import openpyxl
from datetime import datetime

def main():
    try:
        # 現在の年と月を取得
        current_year = datetime.now().year
        current_month = datetime.now().month
        
        # Excelファイルのパスを作成
        excel_path = f"./{current_year}/{current_month:02}.xlsx"
        
        # ファイルが存在するかチェック
        if os.path.exists(excel_path):
            # Excelファイルを開く
            wb = openpyxl.load_workbook(excel_path)
            
            # 当月のシート名を取得
            sheet_name = f"{current_month}月"
            
            # シートが存在するかチェック
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # 当日の日付を取得
                current_day = datetime.now().day
                
                # 新しいワークブックを作成
                new_wb = openpyxl.Workbook()
                new_sheet = new_wb.active
                new_sheet.title = "copy"
                
                # ロジック: 当日の日付と一致する行をコピー
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    date_str = row[2]
                    if date_matches_current_day(date_str, current_day):
                        new_sheet.append(row)
                
                # 新しいファイルを保存
                new_file_path = f"./copy_{current_year}_{current_month:02}.xlsx"
                new_wb.save(new_file_path)
                
                print("コピーが完了しました。")
            else:
                print(f"{sheet_name}シートが存在しません。")
        else:
            print(f"{excel_path}が存在しません。")
    except Exception as e:
        print(f"エラーが発生しました: {e}")

# ロジック: 日付が当日と一致するかどうかを判定
def date_matches_current_day(date_str, current_day):
    try:
        month, day = map(int, date_str.split('/'))
        return month == datetime.now().month and day == current_day
    except:
        return False

if __name__ == "__main__":
    main()
