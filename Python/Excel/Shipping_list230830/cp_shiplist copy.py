import os
import openpyxl
import xlwings as xw
from datetime import datetime, timedelta
import time
from tqdm import tqdm
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Select Days")
        self.selected_days = []

        self.current_day = tk.IntVar()
        self.current_day.set(datetime.now().day)

        self.label = ttk.Label(root, text="Select starting day:")
        self.label.pack(padx=10, pady=5)

        self.day_combobox = ttk.Combobox(root, textvariable=self.current_day, values=self.get_selectable_days())
        self.day_combobox.pack(padx=10, pady=5)

        self.confirm_button = ttk.Button(root, text="Confirm", command=self.confirm_selection)
        self.confirm_button.pack(padx=10, pady=10)

        # Bind the close event to a function
        self.root.protocol("WM_DELETE_WINDOW", self.close_app)

    def get_selectable_days(self):
        today = datetime.now().day
        current_month = datetime.now().month
        current_year = datetime.now().year
        last_day_of_month = (datetime(current_year, current_month + 1, 1) - timedelta(days=1)).day
        selectable_days = [(today + i) % last_day_of_month for i in range(3)]
        return [f"{current_month}/{day}/{current_year}" for day in selectable_days]

    def confirm_selection(self):
        today = datetime.now().day
        start_day = self.current_day.get()
        self.selected_days = [(start_day + i) % 30 for i in range(3) if (start_day + i) % 30 >= today]
        self.root.destroy()

    def close_app(self):
        self.root.destroy()

def main(selected_days):
    try:
        current_year, current_month = get_current_year_and_month()
        excel_path, sheet_name = create_excel_path_and_sheet_name(current_year, current_month)

        if os.path.exists(excel_path):
            app, wb, sheet = open_excel_app_and_workbook(excel_path, sheet_name)
            paste_wb, paste_sheet = open_or_create_paste_workbook()

            copy_matching_rows(sheet, selected_days, paste_sheet)

            save_and_close_paste_workbook(paste_wb)
            quit_excel_app(app)
        else:
            show_error_message(f"{excel_path} does not exist.")
    except Exception as e:
        show_error_message(f"An error occurred: {e}")

def get_current_year_and_month():
    now = datetime.now()
    return now.year, now.month

def create_excel_path_and_sheet_name(year, month):
    excel_path = f"./{year}/{month:02}.xlsx"
    sheet_name = f"{month}月"
    return excel_path, sheet_name

def open_excel_app_and_workbook(excel_path, sheet_name):
    app = xw.App(visible=False)
    wb = app.books.open(excel_path)
    sheet = wb.sheets[sheet_name]
    return app, wb, sheet

def open_or_create_paste_workbook():
    paste_wb_path = './sl.xlsx'
    if os.path.exists(paste_wb_path):
        paste_wb = openpyxl.load_workbook(paste_wb_path)
    else:
        paste_wb = openpyxl.Workbook()
    paste_sheet_name = 'paste'
    if paste_sheet_name in paste_wb.sheetnames:
        paste_sheet = paste_wb[paste_sheet_name]
        paste_sheet.delete_rows(4, paste_sheet.max_row)
    else:
        paste_sheet = paste_wb.create_sheet(paste_sheet_name)
    return paste_wb, paste_sheet

def copy_matching_rows(sheet, selected_days, paste_sheet):
    used_range = sheet.used_range
    num_rows = used_range.last_cell.row
    progress_bar = tqdm(total=num_rows, desc="Copying Rows")

    consecutive_none_count = 0
    max_consecutive_none = 5

    try:
        for i, row in enumerate(used_range.rows, start=1):
            date_str = row[2].value

            if date_str is None:
                consecutive_none_count += 1
            else:
                consecutive_none_count = 0

            if consecutive_none_count >= max_consecutive_none:
                break

            if date_matches_selected_days(date_str, selected_days):
                new_row = process_row(row, sheet)
                paste_sheet.append(new_row)

            progress_bar.update(1)
            time.sleep(0.1)

    except KeyboardInterrupt:
        pass

    progress_bar.close()

def show_error_message(message):
    messagebox.showinfo("Error", message)

def save_and_close_paste_workbook(paste_wb):
    paste_wb_path = './sl.xlsx'
    paste_wb.save(paste_wb_path)

def quit_excel_app(app):
    app.quit()


# Logic: Check if date matches selected days
def date_matches_selected_days(date_str, selected_days):
    try:
        day = map(int, date_str.split('/'))
        return day in selected_days
    except:
        return False

if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()
    main(app.selected_days)
