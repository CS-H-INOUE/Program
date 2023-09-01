import os
import openpyxl
import xlwings as xw
from datetime import datetime, timedelta
import time
from tqdm import tqdm
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Select Days")
        self.selected_days = []

        self.current_day = tk.IntVar()
        self.current_day.set(datetime.now().day)

        self.label = ttk.Label(root, text="Select starting day:")
        self.label.pack(padx=10, pady=5)

        self.day_combobox = ttk.Combobox(root, textvariable=self.current_day, values=self.get_selectable_days())
        self.day_combobox.pack(padx=10, pady=5)

        self.confirm_button = ttk.Button(root, text="Confirm", command=self.confirm_selection)
        self.confirm_button.pack(padx=10, pady=10)

        # Bind the close event to a function
        self.root.protocol("WM_DELETE_WINDOW", self.close_app)

    def get_selectable_days(self):
        today = datetime.now().day
        current_month = datetime.now().month
        current_year = datetime.now().year
        last_day_of_month = (datetime(current_year, current_month + 1, 1) - timedelta(days=1)).day
        selectable_days = [(today + i) % last_day_of_month for i in range(3)]
        return [f"{current_month}/{day}/{current_year}" for day in selectable_days]

    def confirm_selection(self):
        today = datetime.now().day
        start_day = self.current_day.get()
        self.selected_days = [(start_day + i) % 30 for i in range(3) if (start_day + i) % 30 >= today]
        self.root.destroy()

    def close_app(self):
        self.root.destroy()

def main(selected_days):
    try:
        # Get current year and month
        current_year = datetime.now().year
        current_month = datetime.now().month
        
        # Create Excel file path
        excel_path = f"./{current_year}/{current_month:02}.xlsx"
        sheet_name = f"{current_month}月"
        
        # Check if the file exists
        if os.path.exists(excel_path):
            # Open xlwings application
            app = xw.App(visible=False)
            wb = app.books.open(excel_path)
            sheet = wb.sheets[sheet_name]
            
            # Open existing workbook for pasting
            paste_wb_path = './sl.xlsx'
            if os.path.exists(paste_wb_path):
                paste_wb = openpyxl.load_workbook(paste_wb_path)
            else:
                paste_wb = openpyxl.Workbook()
            paste_sheet_name = 'paste'
            if paste_sheet_name in paste_wb.sheetnames:
                paste_sheet = paste_wb[paste_sheet_name]
                paste_sheet.delete_rows(4, paste_sheet.max_row)  # Delete existing data from row 4 onwards
            
            else:
                paste_sheet = paste_wb.create_sheet(paste_sheet_name)
            
            # Logic: Copy rows with matching selected days
            used_range = sheet.used_range
            num_rows = used_range.last_cell.row
            progress_bar = tqdm(total=num_rows, desc="Copying Rows")
            
            consecutive_none_count = 0  # Counter for consecutive None values
            max_consecutive_none = 5     # Maximum allowed consecutive None values
            
            try:
                for i, row in enumerate(used_range.rows, start=1):
                    date_str = row[2].value
                    
                    if date_str is None:
                        consecutive_none_count += 1
                    else:
                        consecutive_none_count = 0
                    
                    if consecutive_none_count >= max_consecutive_none:
                        break
                    
                    if date_matches_selected_days(date_str, selected_days):
                        new_row = []
                        for cell in row:
                            if isinstance(cell.value, str) and cell.value.startswith("="):
                                try:
                                    value = sheet.range(cell.address).value
                                    new_row.append(value)
                                except:
                                    new_row.append(cell.value)
                            else:
                                new_row.append(cell.value)
                        paste_sheet.append(new_row)
                    
                    progress_bar.update(1)
                    time.sleep(0.1)  # Adjusted sleep time
            
            except KeyboardInterrupt:
                pass
            
            progress_bar.close()
            
            # Save the changes to the paste workbook
            paste_wb.save(paste_wb_path)
            
            # Quit xlwings application
            app.quit()
        else:
            messagebox.showinfo("Error", f"{excel_path} does not exist.")
    except Exception as e:
        messagebox.showinfo("Error", f"An error occurred: {e}")

# Logic: Check if date matches selected days
def date_matches_selected_days(date_str, selected_days):
    try:
        month, day = map(int, date_str.split('/'))
        return day in selected_days
    except:
        return False

if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()
    main(app.selected_days)
